"""
FINMIND MACRO & COMMODITY SCRAPER v3.0 — BỔ SUNG 11 NGÀNH CÒN THIẾU
======================================================================
Mục tiêu: Cào thêm ~30,000 dòng dữ liệu cho 11 ngành chưa có, rồi APPEND
           vào 2 file Excel + CSV đã tạo trước đó (v2).

11 ngành bổ sung:
    1. Hóa chất          → XLB (Materials ETF), HO=F (Heating Oil)
    2. Tài nguyên Cơ bản → SLX (Steel ETF), PICK (Mining ETF), KOL (Coal ETF)
    3. Xây dựng & VL     → LBS=F (Lumber), XHB (Homebuilders ETF)
    4. CN & Vận tải      → BDRY (Shipping ETF), IYT (Transportation ETF)
    5. Ô tô & Phụ tùng   → PA=F (Palladium)
    6. Hàng cá nhân      → WOOD (Timber ETF), REMX (Rare Earth ETF)
    7. Y tế              → XLV (Healthcare ETF), IBB (Biotech ETF)
    8. Truyền thông      → XLC (Communication ETF)
    9. Du lịch & GT      → JETS (Airlines ETF)
   10. Bất động sản      → IYR, XLRE, ITB (Real Estate ETFs)
   11. CNTT              → SMH, SOXX, IGV (Semiconductor & Software ETFs)
   + Bổ sung liên ngành  → DBA (Agriculture), URA (Uranium), DX-Y.NYB (USD Index)

Nguồn uy tín:
    - Yahoo Finance (ETFs + Futures) — dữ liệu DAILY 5 năm
    - World Bank Open Data API — chỉ số vĩ mô bổ sung

Đầu ra: APPEND vào 2 file đã có:
    - FinMind_Task6_Macro_Commodity_GENERAL_MARKET.xlsx
    - FinMind_Task6_Macro_Commodity_GENERAL_MARKET.csv
"""

import csv
import json
import os
import sys
import time
import random
from datetime import datetime

import requests

# Ép stdout dùng UTF-8
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# =====================================================================
# CẤU HÌNH: 25 TICKER MỚI CHO 11 NGÀNH CÒN THIẾU
# =====================================================================

NEW_COMMODITY_TICKERS = {

    # ═══════════════════════════════════════════════════════════════
    # 1. HÓA CHẤT (Chemicals)
    # ═══════════════════════════════════════════════════════════════
    "XLB": {
        "name_vi": "Chỉ số ngành Vật liệu & Hóa chất (S&P 500)",
        "name_en": "Materials Select Sector SPDR Fund",
        "unit": "USD",
        "category": "Hóa chất",
        "impact": "Tác động: Ngành hóa chất, phân bón (DGC, DCM, DPM, CSV, LAS)",
    },
    "HO=F": {
        "name_vi": "Dầu sưởi (Heating Oil)",
        "name_en": "NY Harbor ULSD Futures",
        "unit": "USD/gallon",
        "category": "Hóa chất",
        "impact": "Tác động: Chi phí nguyên liệu hóa dầu, nhựa, sợi tổng hợp (GVR, BMP)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 2. TÀI NGUYÊN CƠ BẢN (Basic Resources — Thép, Quặng sắt, Than)
    # ═══════════════════════════════════════════════════════════════
    "SLX": {
        "name_vi": "Chỉ số ngành Thép toàn cầu",
        "name_en": "VanEck Steel ETF",
        "unit": "USD",
        "category": "Tài nguyên cơ bản",
        "impact": "Tác động: Ngành thép VN (HPG, HSG, NKG, TIS, POM, SMC)",
    },
    "PICK": {
        "name_vi": "Chỉ số khai khoáng & quặng sắt toàn cầu",
        "name_en": "iShares MSCI Global Metals & Mining Producers ETF",
        "unit": "USD",
        "category": "Tài nguyên cơ bản",
        "impact": "Tác động: Giá quặng sắt → chi phí đầu vào HPG, biên LN ngành thép",
    },
    "KOL": {
        "name_vi": "Chỉ số ngành Than đá toàn cầu",
        "name_en": "VanEck Coal ETF",
        "unit": "USD",
        "category": "Tài nguyên cơ bản",
        "impact": "Tác động: Ngành than VN (NBC, TC6, MDC, TVD), chi phí nhiệt điện",
    },

    # ═══════════════════════════════════════════════════════════════
    # 3. XÂY DỰNG & VẬT LIỆU (Construction & Materials)
    # ═══════════════════════════════════════════════════════════════
    "LBS=F": {
        "name_vi": "Gỗ xẻ xây dựng (Lumber)",
        "name_en": "Random Length Lumber Futures",
        "unit": "USD/1000 board feet",
        "category": "Xây dựng & Vật liệu",
        "impact": "Tác động: Chi phí xây dựng, nội thất, xuất khẩu gỗ (PTB, GDT, SAV)",
    },
    "XHB": {
        "name_vi": "Chỉ số ngành Xây dựng nhà ở (Mỹ)",
        "name_en": "SPDR S&P Homebuilders ETF",
        "unit": "USD",
        "category": "Xây dựng & Vật liệu",
        "impact": "Tác động: Xu hướng xây dựng toàn cầu → VN (VCG, CTD, HBC, ROS)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 4. HÀNG & DỊCH VỤ CÔNG NGHIỆP (Logistics, Cảng, Vận tải biển)
    # ═══════════════════════════════════════════════════════════════
    "BDRY": {
        "name_vi": "Chỉ số cước vận tải biển (Baltic Dry Index proxy)",
        "name_en": "Breakwave Dry Bulk Shipping ETF",
        "unit": "USD",
        "category": "Hàng & DV Công nghiệp",
        "impact": "Tác động: Cước vận tải biển → ngành cảng (GMD, VSC, PHP, HAH, VOS)",
    },
    "IYT": {
        "name_vi": "Chỉ số ngành Vận tải & Logistics (Mỹ)",
        "name_en": "iShares U.S. Transportation ETF",
        "unit": "USD",
        "category": "Hàng & DV Công nghiệp",
        "impact": "Tác động: Xu hướng logistics toàn cầu → VN (GMD, VTP, VNPost)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 5. Ô TÔ & PHỤ TÙNG (Automobiles & Parts)
    # ═══════════════════════════════════════════════════════════════
    "PA=F": {
        "name_vi": "Bạch kim nhóm Palladium (xúc tác ô tô)",
        "name_en": "Palladium Futures",
        "unit": "USD/ounce",
        "category": "Ô tô & Phụ tùng",
        "impact": "Tác động: Nguyên liệu bộ xúc tác khí thải ô tô → ngành ô tô VN (VEA, SVC, HAX, TMT)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 6. HÀNG CÁ NHÂN & GIA DỤNG (Personal & Household Goods)
    # ═══════════════════════════════════════════════════════════════
    "WOOD": {
        "name_vi": "Chỉ số ngành Gỗ & Lâm nghiệp toàn cầu",
        "name_en": "iShares Global Timber & Forestry ETF",
        "unit": "USD",
        "category": "Hàng cá nhân & Gia dụng",
        "impact": "Tác động: Ngành gỗ xuất khẩu VN (PTB, GDT, ACG, TTF, SAV)",
    },
    "REMX": {
        "name_vi": "Chỉ số Đất hiếm & Kim loại chiến lược",
        "name_en": "VanEck Rare Earth/Strategic Metals ETF",
        "unit": "USD",
        "category": "Hàng cá nhân & Gia dụng",
        "impact": "Tác động: Linh kiện điện tử gia dụng, pin, thiết bị thông minh",
    },

    # ═══════════════════════════════════════════════════════════════
    # 7. Y TẾ (Healthcare & Pharmaceuticals)
    # ═══════════════════════════════════════════════════════════════
    "XLV": {
        "name_vi": "Chỉ số ngành Y tế & Dược phẩm (S&P 500)",
        "name_en": "Health Care Select Sector SPDR Fund",
        "unit": "USD",
        "category": "Y tế",
        "impact": "Tác động: Ngành dược VN (DHG, IMP, DMC, DBD, TRA, OPC)",
    },
    "IBB": {
        "name_vi": "Chỉ số ngành Công nghệ Sinh học",
        "name_en": "iShares Biotechnology ETF",
        "unit": "USD",
        "category": "Y tế",
        "impact": "Tác động: Xu hướng R&D dược phẩm toàn cầu → VN (DHG, IMP)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 8. TRUYỀN THÔNG (Media & Communication)
    # ═══════════════════════════════════════════════════════════════
    "XLC": {
        "name_vi": "Chỉ số ngành Truyền thông & Viễn thông (S&P 500)",
        "name_en": "Communication Services Select Sector SPDR Fund",
        "unit": "USD",
        "category": "Truyền thông",
        "impact": "Tác động: Ngành viễn thông & truyền thông VN (VNM, FOX, VGI, VTV)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 9. DU LỊCH & GIẢI TRÍ (Tourism & Leisure)
    # ═══════════════════════════════════════════════════════════════
    "JETS": {
        "name_vi": "Chỉ số ngành Hàng không toàn cầu",
        "name_en": "U.S. Global Jets ETF",
        "unit": "USD",
        "category": "Du lịch & Giải trí",
        "impact": "Tác động: Ngành hàng không & du lịch VN (VJC, HVN, ACV, VTR, DSE)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 10. BẤT ĐỘNG SẢN (Real Estate)
    # ═══════════════════════════════════════════════════════════════
    "IYR": {
        "name_vi": "Chỉ số BĐS toàn cầu (iShares)",
        "name_en": "iShares U.S. Real Estate ETF",
        "unit": "USD",
        "category": "Bất động sản",
        "impact": "Tác động: Xu hướng BĐS toàn cầu → VN (VHM, VIC, NVL, NLG, KDH, DXG)",
    },
    "XLRE": {
        "name_vi": "Chỉ số BĐS S&P 500",
        "name_en": "Real Estate Select Sector SPDR Fund",
        "unit": "USD",
        "category": "Bất động sản",
        "impact": "Tác động: REITs & quỹ BĐS → dòng vốn vào BĐS VN (VHM, NVL, KDH)",
    },
    "ITB": {
        "name_vi": "Chỉ số ngành Xây dựng nhà ở (iShares)",
        "name_en": "iShares U.S. Home Construction ETF",
        "unit": "USD",
        "category": "Bất động sản",
        "impact": "Tác động: Nhu cầu vật liệu xây dựng, xi măng VN (HT1, BCC, BTS)",
    },

    # ═══════════════════════════════════════════════════════════════
    # 11. CÔNG NGHỆ THÔNG TIN (IT & Semiconductors)
    # ═══════════════════════════════════════════════════════════════
    "SMH": {
        "name_vi": "Chỉ số ngành Bán dẫn (VanEck)",
        "name_en": "VanEck Semiconductor ETF",
        "unit": "USD",
        "category": "Công nghệ thông tin",
        "impact": "Tác động: Ngành CNTT & bán dẫn VN (FPT, CMG, ELC, SAM, ITD)",
    },
    "SOXX": {
        "name_vi": "Chỉ số ngành Bán dẫn (iShares)",
        "name_en": "iShares Semiconductor ETF",
        "unit": "USD",
        "category": "Công nghệ thông tin",
        "impact": "Tác động: Chuỗi cung ứng chip → FDI vào VN (Samsung, Intel, Amkor)",
    },
    "IGV": {
        "name_vi": "Chỉ số ngành Phần mềm & Dịch vụ CNTT",
        "name_en": "iShares Expanded Tech-Software Sector ETF",
        "unit": "USD",
        "category": "Công nghệ thông tin",
        "impact": "Tác động: Outsourcing phần mềm VN (FPT Software, CMC, TMA, KMS)",
    },

    # ═══════════════════════════════════════════════════════════════
    # BỔ SUNG LIÊN NGÀNH (Cross-sector indicators)
    # ═══════════════════════════════════════════════════════════════
    "DBA": {
        "name_vi": "Quỹ chỉ số Nông sản tổng hợp",
        "name_en": "Invesco DB Agriculture Fund",
        "unit": "USD",
        "category": "Nông sản (tổng hợp)",
        "impact": "Tác động: Giá nông sản tổng hợp → ngành nông nghiệp VN (HAG, HNG, LTG)",
    },
    "URA": {
        "name_vi": "Chỉ số ngành Uranium & Năng lượng hạt nhân",
        "name_en": "Global X Uranium ETF",
        "unit": "USD",
        "category": "Hóa chất",
        "impact": "Tác động: Xu hướng năng lượng hạt nhân toàn cầu → Điện lực VN",
    },
    "DX-Y.NYB": {
        "name_vi": "Chỉ số sức mạnh đồng USD (Dollar Index)",
        "name_en": "ICE U.S. Dollar Index",
        "unit": "Điểm",
        "category": "Kinh tế vĩ mô toàn cầu",
        "impact": "Tác động trực tiếp: Tỷ giá USD/VND, dòng vốn ngoại, xuất nhập khẩu VN",
    },
}

# Tần suất: Daily, 5 năm
COMMODITY_PERIOD = "5y"
COMMODITY_INTERVAL = "1d"


# =====================================================================
# CẤU HÌNH: 6 CHỈ SỐ VĨ MÔ MỚI TỪ WORLD BANK
# =====================================================================

NEW_WB_INDICATORS = {
    "ST.INT.ARVL": {
        "name_vi": "Lượng khách du lịch quốc tế đến VN",
        "name_en": "International tourism, number of arrivals",
        "unit": "Lượt khách",
        "category": "Du lịch & Giải trí",
        "impact": "Tác động: Ngành du lịch, hàng không, khách sạn (VJC, HVN, ACV, VTR)",
    },
    "SH.XPD.CHEX.GD.ZS": {
        "name_vi": "Chi tiêu y tế quốc gia (% GDP)",
        "name_en": "Current health expenditure (% of GDP)",
        "unit": "% GDP",
        "category": "Y tế",
        "impact": "Tác động: Quy mô thị trường dược phẩm VN (DHG, IMP, DMC)",
    },
    "NE.GDI.TOTL.ZS": {
        "name_vi": "Tổng đầu tư vốn cố định (% GDP)",
        "name_en": "Gross capital formation (% of GDP)",
        "unit": "% GDP",
        "category": "Xây dựng & Bất động sản",
        "impact": "Tác động: Đầu tư xây dựng, hạ tầng, BĐS (VCG, CTD, VHM, NVL)",
    },
    "TX.VAL.TECH.MF.ZS": {
        "name_vi": "Xuất khẩu công nghệ cao (% xuất khẩu CN)",
        "name_en": "High-technology exports (% of manufactured exports)",
        "unit": "%",
        "category": "Công nghệ thông tin",
        "impact": "Tác động: Năng lực công nghệ VN, FDI công nghệ (FPT, Samsung VN)",
    },
    "NV.IND.TOTL.ZS": {
        "name_vi": "Giá trị gia tăng công nghiệp (% GDP)",
        "name_en": "Industry (including construction), value added (% of GDP)",
        "unit": "% GDP",
        "category": "Hàng & DV Công nghiệp",
        "impact": "Tác động: Sức khỏe ngành công nghiệp VN, sản xuất, chế biến",
    },
    "NV.SRV.TOTL.ZS": {
        "name_vi": "Giá trị gia tăng dịch vụ (% GDP)",
        "name_en": "Services, value added (% of GDP)",
        "unit": "% GDP",
        "category": "Du lịch & Truyền thông",
        "impact": "Tác động: Ngành dịch vụ, du lịch, CNTT, truyền thông VN",
    },
}

MACRO_START_YEAR = 2000
MACRO_END_YEAR = 2025
COUNTRY_CODE = "VN"


# =====================================================================
# THỨ TỰ CỘT CHUẨN (PHẢI KHỚP VỚI FILE CŨ)
# =====================================================================

CSV_FIELDNAMES = [
    "data_type", "ticker", "name_vi", "name_en", "category",
    "unit", "impact_vietnam", "date", "open", "high", "low",
    "close", "volume", "source",
]


# =====================================================================
# HÀM THU THẬP DỮ LIỆU
# =====================================================================

def fetch_world_bank_indicator(indicator_code, country="VN", start_year=2000, end_year=2025):
    """Gọi World Bank API lấy chuỗi thời gian 1 chỉ số kinh tế."""
    url = (
        f"https://api.worldbank.org/v2/country/{country}"
        f"/indicator/{indicator_code}"
        f"?date={start_year}:{end_year}&format=json&per_page=500"
    )
    try:
        response = requests.get(url, timeout=15)
        if response.status_code != 200:
            return []

        data = response.json()
        if not data or len(data) < 2 or not data[1]:
            return []

        records = []
        for entry in data[1]:
            year = entry.get("date")
            value = entry.get("value")
            if value is not None:
                records.append({
                    "year": int(year),
                    "value": round(value, 4) if isinstance(value, float) else value,
                })
        records.sort(key=lambda x: x["year"])
        return records

    except Exception as e:
        print(f"   [!] Ngoại lệ World Bank ({indicator_code}): {e}")
        return []


def collect_new_macro_data():
    """Thu thập 6 chỉ số vĩ mô MỚI từ World Bank."""
    print("=" * 65)
    print("  📊 PHẦN 1: THU THẬP CHỈ SỐ VĨ MÔ BỔ SUNG (6 chỉ số)")
    print(f"  Nguồn: World Bank Open Data API")
    print("=" * 65)

    new_macro_rows = []
    total = len(NEW_WB_INDICATORS)

    for idx, (code, info) in enumerate(NEW_WB_INDICATORS.items(), 1):
        print(f"\n[{idx}/{total}] Đang tải: {info['name_vi']}...")

        time_series = fetch_world_bank_indicator(
            indicator_code=code,
            country=COUNTRY_CODE,
            start_year=MACRO_START_YEAR,
            end_year=MACRO_END_YEAR,
        )

        if time_series:
            print(f"   + Thành công: {len(time_series)} năm ({time_series[0]['year']}-{time_series[-1]['year']})")
        else:
            print(f"   - Không có dữ liệu.")

        for point in time_series:
            new_macro_rows.append({
                "data_type": "Macro",
                "ticker": code,
                "name_vi": info["name_vi"],
                "name_en": info["name_en"],
                "category": info["category"],
                "unit": info["unit"],
                "impact_vietnam": info["impact"],
                "date": f"{point['year']}-01-01",
                "open": "",
                "high": "",
                "low": "",
                "close": point["value"],
                "volume": "",
                "source": "World Bank",
            })

        time.sleep(random.uniform(0.3, 0.7))

    print(f"\n[+] Vĩ mô bổ sung: {len(new_macro_rows)} dòng mới.")
    return new_macro_rows


def collect_new_commodity_data():
    """Thu thập 25 ticker MỚI (ETF + Futures) daily 5 năm từ Yahoo Finance."""
    print("\n" + "=" * 65)
    print("  🌍 PHẦN 2: THU THẬP 25 TICKER MỚI (11 NGÀNH)")
    print(f"  Nguồn: Yahoo Finance — Daily × 5 năm")
    print("=" * 65)

    try:
        import yfinance as yf
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-U", "yfinance"])
        import yfinance as yf

    new_commodity_rows = []
    total = len(NEW_COMMODITY_TICKERS)
    failed_tickers = []

    for idx, (ticker, info) in enumerate(NEW_COMMODITY_TICKERS.items(), 1):
        print(f"\n[{idx}/{total}] Đang tải: {info['name_vi']} ({ticker})...")

        try:
            obj = yf.Ticker(ticker)
            hist = obj.history(period=COMMODITY_PERIOD, interval=COMMODITY_INTERVAL)

            if hist.empty:
                print(f"   [-] Không có dữ liệu cho {ticker}. Bỏ qua.")
                failed_tickers.append(ticker)
                continue

            count = 0
            for date, row in hist.iterrows():
                # Bỏ qua dòng NaN
                if row["Close"] != row["Close"]:
                    continue

                new_commodity_rows.append({
                    "data_type": "Commodity",
                    "ticker": ticker,
                    "name_vi": info["name_vi"],
                    "name_en": info["name_en"],
                    "category": info["category"],
                    "unit": info["unit"],
                    "impact_vietnam": info["impact"],
                    "date": date.strftime("%Y-%m-%d"),
                    "open": round(float(row["Open"]), 4) if row["Open"] == row["Open"] else "",
                    "high": round(float(row["High"]), 4) if row["High"] == row["High"] else "",
                    "low": round(float(row["Low"]), 4) if row["Low"] == row["Low"] else "",
                    "close": round(float(row["Close"]), 4) if row["Close"] == row["Close"] else "",
                    "volume": int(row["Volume"]) if row["Volume"] == row["Volume"] else 0,
                    "source": "Yahoo Finance",
                })
                count += 1

            print(f"   + Thành công: {count} ngày ({hist.index[0].strftime('%Y-%m-%d')} → {hist.index[-1].strftime('%Y-%m-%d')})")

        except Exception as e:
            print(f"   [!] Lỗi khi tải {ticker}: {e}")
            failed_tickers.append(ticker)

        time.sleep(random.uniform(0.3, 0.6))

    if failed_tickers:
        print(f"\n[!] Các ticker bị lỗi/không có dữ liệu: {', '.join(failed_tickers)}")

    print(f"\n[+] Hàng hóa & ETF bổ sung: {len(new_commodity_rows)} dòng mới.")
    return new_commodity_rows


# =====================================================================
# ĐỌC DỮ LIỆU CŨ TỪ FILE CSV
# =====================================================================

def read_existing_csv(filepath):
    """Đọc toàn bộ dữ liệu cũ từ file CSV đã tạo trước."""
    rows = []
    if not os.path.exists(filepath):
        print(f"[!] Không tìm thấy file cũ: {filepath}")
        return rows

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(dict(row))

    print(f"[*] Đã đọc {len(rows):,} dòng dữ liệu cũ từ {filepath}")
    return rows


# =====================================================================
# XUẤT FILE EXCEL + CSV (GHI ĐÈ VỚI DỮ LIỆU GỘP)
# =====================================================================

def export_combined_excel(all_rows, filepath):
    """Tạo file Excel với 3 sheet từ dữ liệu gộp (cũ + mới)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=10)

    col_widths = {
        "data_type": 12, "ticker": 22, "name_vi": 38, "name_en": 50,
        "category": 26, "unit": 18, "impact_vietnam": 60,
        "date": 13, "open": 14, "high": 14, "low": 14,
        "close": 14, "volume": 15, "source": 15,
    }

    def write_sheet(ws, rows_data, sheet_title):
        ws.title = sheet_title

        # Header
        for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows_data, 2):
            for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
                value = row_data.get(field, "")

                # Chuyển đổi số nếu cần
                if field in ("open", "high", "low", "close", "volume") and value != "" and value is not None:
                    try:
                        value = float(value)
                        if field == "volume":
                            value = int(value)
                    except (ValueError, TypeError):
                        pass

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

                if field in ("open", "high", "low", "close") and isinstance(value, (int, float)):
                    cell.number_format = "#,##0.0000"
                    cell.alignment = Alignment(horizontal="right")
                elif field == "volume" and isinstance(value, (int, float)):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")

        # Column widths
        for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 15)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDNAMES))}{len(rows_data) + 1}"

    # Phân loại dữ liệu
    macro_rows = [r for r in all_rows if r.get("data_type") == "Macro"]
    commodity_rows = [r for r in all_rows if r.get("data_type") != "Macro"]

    # Sheet 1: Vĩ mô Việt Nam
    ws_macro = wb.active
    write_sheet(ws_macro, macro_rows, "Vĩ mô Việt Nam")

    # Sheet 2: Hàng hóa & ETF
    ws_commodity = wb.create_sheet()
    write_sheet(ws_commodity, commodity_rows, "Giá Hàng hóa & ETF")

    # Sheet 3: Tổng hợp
    ws_all = wb.create_sheet()
    all_sorted = sorted(all_rows, key=lambda x: x.get("date", ""))
    write_sheet(ws_all, all_sorted, "Tổng hợp (All Data)")

    wb.save(filepath)
    size_mb = round(os.path.getsize(filepath) / (1024 * 1024), 2)
    print(f"[v] Đã xuất file Excel: {filepath} ({size_mb} MB)")


def export_combined_csv(all_rows, filepath):
    """Xuất file CSV phẳng gộp toàn bộ dữ liệu."""
    all_rows.sort(key=lambda x: (x.get("data_type", ""), x.get("ticker", ""), x.get("date", "")))

    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        writer.writerows(all_rows)

    size_mb = round(os.path.getsize(filepath) / (1024 * 1024), 2)
    print(f"[v] Đã xuất file CSV: {filepath} ({size_mb} MB)")


# =====================================================================
# MAIN — ĐỌC CŨ → CÀO MỚI → GỘP → GHI ĐÈ 2 FILE
# =====================================================================

if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  📈 FINMIND SCRAPER v3.0 — BỔ SUNG 11 NGÀNH CÒN THIẾU         ║")
    print("║  Mục tiêu: Thêm ~30K dòng → Gộp vào file cũ (19K)             ║")
    print("║  Nguồn: World Bank API + Yahoo Finance (25 ticker mới)          ║")
    print("╚══════════════════════════════════════════════════════════════════╝\n")

    start_time = time.time()

    XLSX_FILE = "FinMind_Task6_Macro_Commodity_GENERAL_MARKET.xlsx"
    CSV_FILE = "FinMind_Task6_Macro_Commodity_GENERAL_MARKET.csv"

    # ──── BƯỚC 1: Đọc dữ liệu cũ ────
    print("=" * 65)
    print("  📂 BƯỚC 1: ĐỌC DỮ LIỆU CŨ")
    print("=" * 65)
    old_rows = read_existing_csv(CSV_FILE)

    # ──── BƯỚC 2: Cào dữ liệu mới ────
    new_macro_rows = collect_new_macro_data()
    new_commodity_rows = collect_new_commodity_data()

    all_new_rows = new_macro_rows + new_commodity_rows
    print(f"\n[+] TỔNG DỮ LIỆU MỚI: {len(all_new_rows):,} dòng")

    # ──── BƯỚC 3: Gộp cũ + mới ────
    combined_rows = old_rows + all_new_rows
    total_combined = len(combined_rows)

    print(f"\n{'=' * 65}")
    print(f"  📦 BƯỚC 3: GỘP VÀ XUẤT FILE")
    print(f"  Cũ: {len(old_rows):,} + Mới: {len(all_new_rows):,} = Tổng: {total_combined:,}")
    print(f"{'=' * 65}")

    # Lưu vào file tạm trước để tránh lỗi PermissionError khi file gốc đang mở
    XLSX_TEMP = XLSX_FILE + ".tmp.xlsx"
    CSV_TEMP = CSV_FILE + ".tmp.csv"

    export_combined_excel(combined_rows, XLSX_TEMP)
    export_combined_csv(combined_rows, CSV_TEMP)

    # Thay thế file gốc bằng file tạm
    import shutil
    for temp_f, orig_f in [(XLSX_TEMP, XLSX_FILE), (CSV_TEMP, CSV_FILE)]:
        try:
            if os.path.exists(orig_f):
                os.replace(temp_f, orig_f)
            else:
                shutil.move(temp_f, orig_f)
            print(f"[v] Đã cập nhật: {orig_f}")
        except PermissionError:
            # File gốc đang mở → giữ file tạm, thông báo user
            final_name = orig_f.replace("GENERAL_MARKET", "GENERAL_MARKET_v3")
            shutil.move(temp_f, final_name)
            print(f"[!] File '{orig_f}' đang mở. Đã lưu bản mới: {final_name}")

    # ──── TỔNG KẾT ────
    elapsed = round(time.time() - start_time, 1)

    # Thống kê theo ngành
    category_stats = {}
    for row in combined_rows:
        cat = row.get("category", "Khác")
        category_stats[cat] = category_stats.get(cat, 0) + 1

    print(f"\n{'=' * 65}")
    print(f"  ✅ HOÀN TẤT BỔ SUNG 11 NGÀNH!")
    print(f"{'=' * 65}")
    print(f"  📊 Dữ liệu cũ (v2)    : {len(old_rows):,} dòng")
    print(f"  ➕ Dữ liệu mới (v3)    : {len(all_new_rows):,} dòng")
    print(f"  📋 TỔNG CỘNG GỘP       : {total_combined:,} dòng")
    print(f"  ⏱️  Thời gian cào       : {elapsed} giây")
    print(f"\n  📁 File đã cập nhật:")
    print(f"     1. {XLSX_FILE}")
    print(f"     2. {CSV_FILE}")
    print(f"\n  📊 Phân bổ theo ngành:")
    for cat, count in sorted(category_stats.items(), key=lambda x: -x[1]):
        print(f"     • {cat:40s} : {count:>6,} dòng")
    print(f"{'=' * 65}")

