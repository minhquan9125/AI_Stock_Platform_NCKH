"""
FINMIND MACRO & COMMODITY SCRAPER v2.0 — BẢN MỞ RỘNG 15K-20K DÒNG
=====================================================================
Task 6: Dữ liệu Kinh tế Vĩ mô & Giá Hàng hóa (Macroeconomic & Commodity Data)

Thay đổi so với v1.0:
    - Tần suất hàng hóa: Monthly → Daily (x50 lần dữ liệu)
    - Thêm 3 hàng hóa: Lúa mì, Gạo thô, Bạch kim (tổng 15 loại)
    - Xuất 2 file duy nhất: .xlsx (2 sheet, cho nhóm đọc) + .csv (flat, cho train model)
    - Mục tiêu: 15,000 - 20,000 dòng dữ liệu

Nguồn uy tín:
    1. World Bank Open Data API (api.worldbank.org)
    2. Yahoo Finance (finance.yahoo.com)

Đầu ra:
    - FinMind_Task6_Macro_Commodity_GENERAL_MARKET.xlsx  (2 sheet)
    - FinMind_Task6_Macro_Commodity_GENERAL_MARKET.csv   (bảng phẳng gộp)
"""

import csv
import json
import os
import sys
import time
import random
from datetime import datetime

import requests

# Ép stdout dùng UTF-8: tránh UnicodeEncodeError khi in tiếng Việt có dấu
# trên terminal Windows mặc định dùng codepage cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# =====================================================================
# PHẦN 1: CẤU HÌNH CHỈ SỐ VĨ MÔ VIỆT NAM (WORLD BANK API)
# =====================================================================

WORLD_BANK_INDICATORS = {
    "NY.GDP.MKTP.KD.ZG": {
        "name_vi": "Tăng trưởng GDP",
        "name_en": "GDP Growth (annual %)",
        "unit": "%",
    },
    "NY.GDP.MKTP.CD": {
        "name_vi": "GDP (USD)",
        "name_en": "GDP (current US$)",
        "unit": "USD",
    },
    "FP.CPI.TOTL.ZG": {
        "name_vi": "Chỉ số CPI (Lạm phát)",
        "name_en": "Inflation, consumer prices (annual %)",
        "unit": "%",
    },
    "FR.INR.LEND": {
        "name_vi": "Lãi suất cho vay",
        "name_en": "Lending interest rate (%)",
        "unit": "%",
    },
    "FR.INR.DPST": {
        "name_vi": "Lãi suất tiền gửi",
        "name_en": "Deposit interest rate (%)",
        "unit": "%",
    },
    "PA.NUS.FCRF": {
        "name_vi": "Tỷ giá USD/VND (bình quân năm)",
        "name_en": "Official exchange rate (LCU per US$, period average)",
        "unit": "VND/USD",
    },
    "FD.AST.PRVT.GD.ZS": {
        "name_vi": "Tín dụng khu vực tư nhân (% GDP)",
        "name_en": "Domestic credit to private sector (% of GDP)",
        "unit": "% GDP",
    },
    "FS.AST.DOMS.GD.ZS": {
        "name_vi": "Tín dụng nội địa ngân hàng (% GDP)",
        "name_en": "Domestic credit provided by financial sector (% of GDP)",
        "unit": "% GDP",
    },
    "FI.RES.TOTL.CD": {
        "name_vi": "Dự trữ ngoại hối (bao gồm vàng)",
        "name_en": "Total reserves (includes gold, current US$)",
        "unit": "USD",
    },
    "NE.RSB.GNFS.ZS": {
        "name_vi": "Cán cân thương mại (% GDP)",
        "name_en": "External balance on goods and services (% of GDP)",
        "unit": "% GDP",
    },
    "BX.KLT.DINV.WD.GD.ZS": {
        "name_vi": "Vốn FDI ròng chảy vào (% GDP)",
        "name_en": "Foreign direct investment, net inflows (% of GDP)",
        "unit": "% GDP",
    },
}

MACRO_START_YEAR = 2000
MACRO_END_YEAR = 2025
COUNTRY_CODE = "VN"


# =====================================================================
# PHẦN 2: CẤU HÌNH HÀNG HÓA THẾ GIỚI (YAHOO FINANCE) — 15 LOẠI
# =====================================================================

COMMODITY_TICKERS = {
    # === DẦU & NĂNG LƯỢNG ===
    "BZ=F": {
        "name_vi": "Dầu thô Brent",
        "name_en": "Brent Crude Oil Futures",
        "unit": "USD/thùng",
        "category": "Năng lượng",
        "impact": "Tác động: Dầu khí (PVD, PVS, PLX), vận tải, lạm phát",
    },
    "CL=F": {
        "name_vi": "Dầu thô WTI",
        "name_en": "WTI Crude Oil Futures",
        "unit": "USD/thùng",
        "category": "Năng lượng",
        "impact": "Tác động: Giá xăng dầu, chi phí vận tải, lạm phát",
    },
    "NG=F": {
        "name_vi": "Khí tự nhiên",
        "name_en": "Natural Gas Futures",
        "unit": "USD/MMBtu",
        "category": "Năng lượng",
        "impact": "Tác động: Chi phí năng lượng, ngành điện (POW, PPC)",
    },
    # === KIM LOẠI QUÝ ===
    "GC=F": {
        "name_vi": "Vàng",
        "name_en": "Gold Futures",
        "unit": "USD/ounce",
        "category": "Kim loại quý",
        "impact": "Tác động: Tâm lý trú ẩn an toàn, lạm phát, tỷ giá",
    },
    "SI=F": {
        "name_vi": "Bạc",
        "name_en": "Silver Futures",
        "unit": "USD/ounce",
        "category": "Kim loại quý",
        "impact": "Tác động: Ngành công nghệ, điện tử, kim loại quý",
    },
    "PL=F": {
        "name_vi": "Bạch kim",
        "name_en": "Platinum Futures",
        "unit": "USD/ounce",
        "category": "Kim loại quý",
        "impact": "Tác động: Ngành ô tô (xúc tác khí thải), trang sức cao cấp",
    },
    # === KIM LOẠI CÔNG NGHIỆP ===
    "HG=F": {
        "name_vi": "Đồng",
        "name_en": "Copper Futures",
        "unit": "USD/pound",
        "category": "Kim loại công nghiệp",
        "impact": "Tác động: Chỉ báo sức khỏe kinh tế, ngành xây dựng, điện",
    },
    # === NÔNG SẢN ===
    "ZC=F": {
        "name_vi": "Ngô",
        "name_en": "Corn Futures",
        "unit": "Cent/bushel",
        "category": "Nông sản",
        "impact": "Tác động: Ngành chăn nuôi, thức ăn gia súc (DBC, Dabaco)",
    },
    "ZS=F": {
        "name_vi": "Đậu tương",
        "name_en": "Soybean Futures",
        "unit": "Cent/bushel",
        "category": "Nông sản",
        "impact": "Tác động: Ngành thực phẩm, dầu ăn, chăn nuôi",
    },
    "ZW=F": {
        "name_vi": "Lúa mì",
        "name_en": "Wheat Futures",
        "unit": "Cent/bushel",
        "category": "Nông sản",
        "impact": "Tác động: Giá lương thực, an ninh lương thực, lạm phát thực phẩm",
    },
    "CT=F": {
        "name_vi": "Bông vải (Cotton)",
        "name_en": "Cotton Futures",
        "unit": "Cent/pound",
        "category": "Nông sản",
        "impact": "Tác động: Ngành dệt may (TCM, STK, TNG, MSH)",
    },
    "KC=F": {
        "name_vi": "Cà phê Arabica",
        "name_en": "Coffee C Futures",
        "unit": "Cent/pound",
        "category": "Nông sản",
        "impact": "Tác động: VN là top 2 xuất khẩu cà phê thế giới (Robusta)",
    },
    "SB=F": {
        "name_vi": "Đường thô",
        "name_en": "Sugar #11 Futures",
        "unit": "Cent/pound",
        "category": "Nông sản",
        "impact": "Tác động: Ngành mía đường (SBT, QNS, LSS, KTS)",
    },
    "RB=F": {
        "name_vi": "Cao su tự nhiên",
        "name_en": "Rubber Futures",
        "unit": "Cent/pound",
        "category": "Nông sản",
        "impact": "Tác động: Ngành cao su (GVR, DPR, PHR, TRC)",
    },
    "ZR=F": {
        "name_vi": "Gạo thô",
        "name_en": "Rough Rice Futures",
        "unit": "Cent/cwt",
        "category": "Nông sản",
        "impact": "Tác động: VN là top 3 xuất khẩu gạo thế giới (Lộc Trời, Vinaseed)",
    },
}

# Cấu hình lấy dữ liệu DAILY trong 5 năm gần nhất
COMMODITY_PERIOD = "5y"
COMMODITY_INTERVAL = "1d"  # ĐỔI TỪ "1mo" SANG "1d" (DAILY)


# =====================================================================
# HÀM THU THẬP DỮ LIỆU
# =====================================================================

def fetch_world_bank_indicator(indicator_code, country="VN", start_year=2000, end_year=2025):
    """Gọi World Bank API lấy chuỗi thời gian 1 chỉ số kinh tế."""
    url = (
        f"https://api.worldbank.org/v2/country/{country}"
        f"/indicator/{indicator_code}"
        f"?date={start_year}:{end_year}&format=json&per_page=500"
    )
    try:
        response = requests.get(url, timeout=15)
        if response.status_code != 200:
            print(f"   [!] Lỗi HTTP {response.status_code} cho {indicator_code}")
            return []

        data = response.json()
        if not data or len(data) < 2 or not data[1]:
            print(f"   [-] Không có dữ liệu World Bank cho {indicator_code}")
            return []

        records = []
        for entry in data[1]:
            year = entry.get("date")
            value = entry.get("value")
            if value is not None:
                records.append({
                    "year": int(year),
                    "value": round(value, 4) if isinstance(value, float) else value,
                })
        records.sort(key=lambda x: x["year"])
        return records

    except Exception as e:
        print(f"   [!] Ngoại lệ World Bank ({indicator_code}): {e}")
        return []


def collect_vietnam_macro_data():
    """Thu thập toàn bộ chỉ số vĩ mô Việt Nam từ World Bank."""
    print("=" * 65)
    print("  📊 PHẦN 1: THU THẬP DỮ LIỆU KINH TẾ VĨ MÔ VIỆT NAM")
    print(f"  Nguồn: World Bank Open Data API (api.worldbank.org)")
    print(f"  Phạm vi: {MACRO_START_YEAR} - {MACRO_END_YEAR}")
    print("=" * 65)

    macro_rows = []
    total = len(WORLD_BANK_INDICATORS)

    for idx, (code, info) in enumerate(WORLD_BANK_INDICATORS.items(), 1):
        print(f"\n[{idx}/{total}] Đang tải: {info['name_vi']} ({info['name_en']})...")

        time_series = fetch_world_bank_indicator(
            indicator_code=code,
            country=COUNTRY_CODE,
            start_year=MACRO_START_YEAR,
            end_year=MACRO_END_YEAR,
        )

        if time_series:
            print(f"   + Thành công: {len(time_series)} năm ({time_series[0]['year']}-{time_series[-1]['year']})")
        else:
            print(f"   - Không có dữ liệu.")

        for point in time_series:
            macro_rows.append({
                "data_type": "Macro",
                "ticker": code,
                "name_vi": info["name_vi"],
                "name_en": info["name_en"],
                "category": "Kinh tế vĩ mô Việt Nam",
                "unit": info["unit"],
                "impact_vietnam": "Tác động trực tiếp lên lãi suất, tỷ giá, dòng vốn TTCK",
                "date": f"{point['year']}-01-01",
                "open": None,
                "high": None,
                "low": None,
                "close": point["value"],
                "volume": None,
                "source": "World Bank",
            })

        time.sleep(random.uniform(0.3, 0.7))

    print(f"\n[+] Vĩ mô VN: Tổng cộng {len(macro_rows)} dòng dữ liệu.")
    return macro_rows


def collect_commodity_data():
    """Thu thập giá hàng hóa DAILY 5 năm từ Yahoo Finance."""
    print("\n" + "=" * 65)
    print("  🌍 PHẦN 2: THU THẬP GIÁ HÀNG HÓA THẾ GIỚI (DAILY)")
    print(f"  Nguồn: Yahoo Finance (finance.yahoo.com)")
    print(f"  Phạm vi: {COMMODITY_PERIOD}, tần suất: {COMMODITY_INTERVAL}")
    print(f"  Số loại hàng hóa: {len(COMMODITY_TICKERS)}")
    print("=" * 65)

    try:
        import yfinance as yf
    except ImportError:
        print("\n[!] Đang cài yfinance...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-U", "yfinance"])
        import yfinance as yf

    commodity_rows = []
    total = len(COMMODITY_TICKERS)

    for idx, (ticker, info) in enumerate(COMMODITY_TICKERS.items(), 1):
        print(f"\n[{idx}/{total}] Đang tải: {info['name_vi']} ({ticker})...")

        try:
            obj = yf.Ticker(ticker)
            hist = obj.history(period=COMMODITY_PERIOD, interval=COMMODITY_INTERVAL)

            if hist.empty:
                print(f"   [-] Không có dữ liệu cho {ticker}.")
                continue

            count = 0
            for date, row in hist.iterrows():
                # Bỏ qua dòng NaN toàn bộ
                if row["Close"] != row["Close"]:
                    continue

                commodity_rows.append({
                    "data_type": "Commodity",
                    "ticker": ticker,
                    "name_vi": info["name_vi"],
                    "name_en": info["name_en"],
                    "category": info["category"],
                    "unit": info["unit"],
                    "impact_vietnam": info["impact"],
                    "date": date.strftime("%Y-%m-%d"),
                    "open": round(float(row["Open"]), 4) if row["Open"] == row["Open"] else None,
                    "high": round(float(row["High"]), 4) if row["High"] == row["High"] else None,
                    "low": round(float(row["Low"]), 4) if row["Low"] == row["Low"] else None,
                    "close": round(float(row["Close"]), 4) if row["Close"] == row["Close"] else None,
                    "volume": int(row["Volume"]) if row["Volume"] == row["Volume"] else 0,
                    "source": "Yahoo Finance",
                })
                count += 1

            print(f"   + Thành công: {count} ngày giao dịch ({hist.index[0].strftime('%Y-%m-%d')} → {hist.index[-1].strftime('%Y-%m-%d')})")

        except Exception as e:
            print(f"   [!] Lỗi khi tải {ticker}: {e}")

        time.sleep(random.uniform(0.3, 0.6))

    print(f"\n[+] Hàng hóa TG: Tổng cộng {len(commodity_rows)} dòng dữ liệu.")
    return commodity_rows


# =====================================================================
# XUẤT FILE EXCEL (.xlsx) VÀ CSV
# =====================================================================

# Thứ tự cột chuẩn cho cả 2 file output
CSV_FIELDNAMES = [
    "data_type", "ticker", "name_vi", "name_en", "category",
    "unit", "impact_vietnam", "date", "open", "high", "low",
    "close", "volume", "source",
]


def export_excel(macro_rows, commodity_rows, filepath):
    """
    Xuất file Excel (.xlsx) với 2 sheet riêng biệt:
        Sheet 1: "Vĩ mô Việt Nam"  — Cho nhóm đọc & kiểm tra nhanh
        Sheet 2: "Giá Hàng hóa TG" — Dữ liệu daily commodities
    """
    try:
        import openpyxl
    except ImportError:
        print("[!] Đang cài openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-U", "openpyxl"])
        import openpyxl

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ─── TIỆN ÍCH: Định dạng sheet đẹp ───
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=10)
    number_format_4dp = "#,##0.0000"
    number_format_int = "#,##0"

    def style_sheet(ws, rows_data, sheet_title):
        """Tạo và định dạng 1 sheet Excel chuẩn chuyên nghiệp."""
        ws.title = sheet_title

        # Viết header
        for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Viết dữ liệu
        for row_idx, row_data in enumerate(rows_data, 2):
            for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
                value = row_data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border

                # Căn phải cho cột số
                if field in ("open", "high", "low", "close") and value is not None:
                    cell.number_format = number_format_4dp
                    cell.alignment = Alignment(horizontal="right")
                elif field == "volume" and value is not None:
                    cell.number_format = number_format_int
                    cell.alignment = Alignment(horizontal="right")

        # Auto-fit độ rộng cột (ước lượng)
        col_widths = {
            "data_type": 12, "ticker": 22, "name_vi": 30, "name_en": 45,
            "category": 24, "unit": 14, "impact_vietnam": 55,
            "date": 13, "open": 14, "high": 14, "low": 14,
            "close": 14, "volume": 15, "source": 15,
        }
        for col_idx, field in enumerate(CSV_FIELDNAMES, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 15)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDNAMES))}{len(rows_data) + 1}"

    # Sheet 1: Vĩ mô
    ws_macro = wb.active
    style_sheet(ws_macro, macro_rows, "Vĩ mô Việt Nam")

    # Sheet 2: Hàng hóa
    ws_commodity = wb.create_sheet()
    style_sheet(ws_commodity, commodity_rows, "Giá Hàng hóa TG")

    # Sheet 3: Tổng hợp (merged, cho tiện)
    ws_all = wb.create_sheet()
    all_rows = macro_rows + commodity_rows
    all_rows.sort(key=lambda x: x["date"])
    style_sheet(ws_all, all_rows, "Tổng hợp (All Data)")

    wb.save(filepath)
    size_kb = round(os.path.getsize(filepath) / 1024, 1)
    print(f"[v] Đã xuất file Excel: {filepath} ({size_kb} KB)")


def export_csv_flat(all_rows, filepath):
    """Xuất file CSV phẳng (flat table) gộp cả Macro + Commodity cho train model."""
    all_rows.sort(key=lambda x: (x["data_type"], x["ticker"], x["date"]))

    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        writer.writerows(all_rows)

    size_kb = round(os.path.getsize(filepath) / 1024, 1)
    print(f"[v] Đã xuất file CSV: {filepath} ({size_kb} KB)")


# =====================================================================
# KHU VỰC THỰC THI CHÍNH
# =====================================================================

if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  📈 FINMIND MACRO & COMMODITY SCRAPER v2.0 — BẢN MỞ RỘNG      ║")
    print("║  Task 6: Kinh tế Vĩ mô & Giá Hàng hóa (15K-20K dòng)         ║")
    print("║  Nguồn: World Bank API + Yahoo Finance (DAILY)                  ║")
    print("║  Đầu ra: 1 file Excel (2 sheet) + 1 file CSV (flat)             ║")
    print("╚══════════════════════════════════════════════════════════════════╝\n")

    start_time = time.time()

    # ──── PHẦN 1: Vĩ mô Việt Nam ────
    macro_rows = collect_vietnam_macro_data()

    # ──── PHẦN 2: Hàng hóa Thế giới (DAILY) ────
    commodity_rows = collect_commodity_data()

    # ──── TỔNG HỢP & XUẤT FILE ────
    all_rows = macro_rows + commodity_rows
    total_rows = len(all_rows)

    print("\n" + "=" * 65)
    print("  📦 ĐANG XUẤT FILE...")
    print("=" * 65)

    xlsx_file = "FinMind_Task6_Macro_Commodity_GENERAL_MARKET.xlsx"
    csv_file = "FinMind_Task6_Macro_Commodity_GENERAL_MARKET.csv"

    export_excel(macro_rows, commodity_rows, xlsx_file)
    export_csv_flat(all_rows, csv_file)

    # ──── BÁO CÁO TỔNG KẾT ────
    elapsed = round(time.time() - start_time, 1)

    print("\n" + "=" * 65)
    print("  ✅ HOÀN TẤT THU THẬP DỮ LIỆU TASK 6 — BẢN MỞ RỘNG!")
    print("=" * 65)
    print(f"  📊 Vĩ mô Việt Nam  : {len(macro_rows):,} dòng ({len(WORLD_BANK_INDICATORS)} chỉ số)")
    print(f"  🌍 Hàng hóa TG     : {len(commodity_rows):,} dòng ({len(COMMODITY_TICKERS)} loại × daily)")
    print(f"  📋 TỔNG CỘNG       : {total_rows:,} dòng dữ liệu")
    print(f"  ⏱️  Thời gian       : {elapsed} giây")
    print(f"\n  📁 File đã xuất (CHỈ 2 FILE):")
    print(f"     1. {xlsx_file}  ← Cho nhóm đọc & kiểm tra")
    print(f"     2. {csv_file}   ← Cho train model AI")
    print(f"\n  💡 Chạy Watchdog để tự động phân loại vào Data Lake.")
    print("=" * 65)
